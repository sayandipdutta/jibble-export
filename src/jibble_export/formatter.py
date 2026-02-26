from pathlib import Path
from uuid import UUID
import logging
from itertools import chain

import pandas as pd
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter


colorfills = {
    "Casual Leave": PatternFill(
        start_color="DDBBDDFF",
        end_color="DDBBDDFF",
        fill_type="solid",
    ),
    "Sick Leave": PatternFill(
        start_color="DD9BBDDF",
        end_color="BBBBDDFF",
        fill_type="solid",
    ),
    "Unpaid Leave": PatternFill(
        start_color="FFFF2222",
        end_color="FFFF2222",
        fill_type="solid",
    ),
    "Present": PatternFill(
        start_color="FFAAFFEF",
        end_color="FFAAFFEF",
        fill_type="solid",
    ),
    "Holidays": PatternFill(
        start_color="FEDCBAFF",
        end_color="FEDCBAFF",
        fill_type="solid",
    ),
    "W/Off": PatternFill(
        start_color="770088FF",
        end_color="770088FF",
        fill_type="solid",
    ),
}


def export_attendance_report(
    tracked_time_report: pd.DataFrame,
    holidays: pd.Series,
    timeoffs: pd.DataFrame,
    id_person_map: dict[UUID, str],
    filename: str,
):
    attendance_report = pd.DataFrame(
        index=tracked_time_report.index,
        columns=tracked_time_report.columns,
        dtype=object,
    )
    notnull_holidays = holidays.dropna()
    for person_id in timeoffs.columns:
        attendance_report.loc[notnull_holidays.index, person_id] = notnull_holidays
    attendance_report[tracked_time_report.notnull()] = "Present"
    attendance_report.loc[attendance_report.index.weekday >= 5, :] = "W/Off"  # ty: ignore[unresolved-attribute]
    for person_id in timeoffs.columns:
        notnull_timeoffs = timeoffs[person_id].dropna()
        attendance_report.loc[notnull_timeoffs.index, person_id] = notnull_timeoffs
    attendance_report = attendance_report.T

    longest_name_chars = max(map(len, id_person_map.values()), default=2)

    Path(filename).parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(filename, engine="openpyxl") as writer:
        attendance_report = attendance_report.reindex(
            sorted(attendance_report.index, key=id_person_map.get)  # ty: ignore[no-matching-overload]
        )
        attendance_report.to_excel(writer, startrow=1, startcol=1, index=False)
        worksheet = writer.sheets["Sheet1"]
        worksheet.sheet_format.defaultColWidth = 15
        dates = attendance_report.columns

        for i, person_id in enumerate(attendance_report.index, start=3):
            worksheet[f"A{i}"] = id_person_map.get(person_id, "N/A")

        worksheet.column_dimensions["A"].width = longest_name_chars + 2
        for col_idx, date in enumerate(
            dates, start=2
        ):  # start=2 because col A is index
            col_letter = get_column_letter(col_idx)
            for i, entry in enumerate(attendance_report[date], start=3):
                if pd.isna(entry):
                    continue
                elif entry.endswith("Casual Leave"):
                    worksheet[f"{col_letter}{i}"].fill = colorfills["Casual Leave"]
                elif entry.endswith("Sick Leave"):
                    worksheet[f"{col_letter}{i}"].fill = colorfills["Sick Leave"]
                elif entry.endswith("Unpaid Leave"):
                    worksheet[f"{col_letter}{i}"].fill = colorfills["Unpaid Leave"]
                elif entry == "Present":
                    worksheet[f"{col_letter}{i}"].fill = colorfills["Present"]
                elif entry == "W/Off":
                    worksheet[f"{col_letter}{i}"].fill = colorfills["W/Off"]
                else:
                    worksheet[f"{col_letter}{i}"].fill = colorfills["Holidays"]

            worksheet[f"{col_letter}1"] = date.strftime("%a")  # weekday name
            worksheet[f"{col_letter}2"] = date.strftime("%d")  # day number

            if date.weekday() >= 5:
                worksheet.column_dimensions[col_letter].width = 6

        start = len(id_person_map) + 4
        worksheet[f"A{start}"] = "Working Days:"
        worksheet[f"B{start}"] = f"{len(attendance_report.columns)}"
        for i, guid in enumerate(attendance_report.index, start=start + 1):
            worksheet[f"A{i}"] = id_person_map[guid]
            worksheet[f"B{i}"] = (attendance_report.loc[guid, :] == "Present").sum()
        worksheet[f"F{start}"] = "Color"
        worksheet[f"G{start}"] = "Represents"
        for i, (kind, value) in enumerate(colorfills.items(), start=start + 1):
            worksheet[f"F{i}"].fill = value
            worksheet[f"G{i}"] = kind
        for cell in chain(worksheet[1], worksheet[2], worksheet[start]):
            cell.font = Font(bold=True)

    logging.info(f"Report successfully exported to {Path(filename).resolve()}")
